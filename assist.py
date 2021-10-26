import re
import openpyxl
import requests
from bs4 import BeautifulSoup
import os
from pathlib import Path

path=Path(os.path.expanduser("~")) / 'Desktop'
workbook =openpyxl.load_workbook(str(path))        #读取表单
print("made by 车")
#修改链接
network = input("请输入青年大学习参学链接: ")
network_1=network.replace('%25','%')
network_2=network_1.replace('#/pages/class/ranking/ranking','api/peopleRankStage')

#读取表单
table1 = workbook.worksheets[0]
table2 = workbook.worksheets[1]
table3 = workbook.worksheets[2]
table4 = workbook.worksheets[3]

rows=table2.max_row            #读取行数
cols=table2.max_column         #读取列数
people=0

def writename(str):
    network_3=network_2+'&level4='+str
    r = requests.get(network_3)
    r.encoding = 'utf-8'
    soup=BeautifulSoup(r.text,'html.parser')                #班级名单的源码
    soup1=soup.encode('utf8').decode('unicode-escape')      #源码编译过后的码
    patternname='(?<=username":")[\u4e00-\u9fa5]+'
    patterndate='\d{4}-\d{1,2}-\d{1,2}'
    namelist=re.findall(patternname,soup1)                  #筛选出名单列表
    datelist=re.findall(patterndate,soup1)                  #筛选出参学日期列表
    
    rows1=table3.max_row            #读取行数
    cols1=table3.max_column         #读取列数
    i=rows1;
    for j in range(0,len(namelist)):
        #print(table2.cell(1,1).value)
        if str=="智能科学技术创新实验2018级团支部":
            table3.cell(row=i+j+1,column=2).value="智能科学与技术2018级1班团支部"
        else:
            table3.cell(row=i+j+1,column=2).value=str
        table3.cell(row=i+j+1,column=3).value=namelist[j]
        table3.cell(row=i+j+1,column=5).value=datelist[j]
    people = len(namelist)       
    del namelist
    del datelist
    return people
    
for i in range(2,rows):
    classes = table2.cell(i,2).value
    table2.cell(row=i,column=4).value=writename(classes)
aa=writename("智能科学技术创新实验2018级团支部")+table2.cell(12,4).value
table2.cell(row=12,column=4).value=aa

for x in range(table4.max_row):
    flag=0
    for y in range(table3.max_row):
        if (table4.cell(x+2,3).value==table3.cell(y+2,3).value)&(table4.cell(x+2,2).value==table3.cell(y+2,2).value):
            flag=1
    if flag==0:
        table1.cell(row=table1.max_row+1,column=2).value=table4.cell(x+2,2).value
        table1.cell(row=table1.max_row,column=3).value=table4.cell(x+2,3).value
        table1.cell(row=table1.max_row,column=4).value=table4.cell(x+2,4).value       
workbook.save(r"C:\Users\che\Desktop\ceshi1.xlsx")     #保存文件
workbook.close()
